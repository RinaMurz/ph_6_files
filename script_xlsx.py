from openpyxl import load_workbook

workbook = load_workbook("tmp/file_example_XLSX_50.xlsx")
sheet = workbook.active
print(sheet.cell(row=1, column=2).value)

for row in sheet.iter_rows(): # получаем все строки и печатаем их
    print(row)

for column in sheet.iter_cols(): # получаем все колонки и печатаем их
    print(column)

for row in sheet.iter_rows():  # печатаем содержимое всех ячеек
    for cell in row:
        print(cell.value)